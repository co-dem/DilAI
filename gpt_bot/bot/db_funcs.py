from datetime import datetime, timezone, timedelta

from configGPT import DB_PATH

import openpyxl
import os


if 'userdb.xlsx' not in os.listdir('src'):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh['A1'] = 'userid'
    sh['B1'] = 'username'
    sh['C1'] = 'payment date'
    wb.save(DB_PATH)

'''
serach_xl can get you number of the first
empty row if you'll set target = None
or
this function can also get you row number
of the specified userid you'll enter in target
'''
def search_xl(target) -> int:
    wb = openpyxl.load_workbook(DB_PATH)
    sh = wb.active

    cr = sh['A1:A500']
    counter = 1

    for i in cr:
        if str(target) == str(i[0].value): 
            return counter
        counter += 1

    wb.close()

def insertNewPaidUser(uid, uname):
    wb = openpyxl.load_workbook(DB_PATH)
    sh = wb.active

    msk_timezone = timezone(timedelta(hours=5))
    dt = str(datetime.now(msk_timezone))

    insert_data = {'A': uid, 'B': uname, 'C': dt[:dt.index('.')]}

    for i in insert_data:
        sh[f"{i}{search_xl(None)}"] = insert_data[i]

    wb.save(DB_PATH)
    wb.close()

def deleteUser(uid):
    wb = openpyxl.load_workbook(DB_PATH)
    sh = wb.active

    cn = search_xl(uid)#? cn - Cell Number

    sh[f'A{cn}'], sh[f'B{cn}'], sh[f'C{cn}'] = '', '', ''

    wb.save(DB_PATH)
    wb.close()

def restore_data(userdata: dict):
    wb = openpyxl.load_workbook(DB_PATH)
    sh = wb.active
    
    for i in range(2, search_xl(None)+1):
        sample = {
            sh[f'A{i}'].value:{
                'lang': None,
                'freeqs': True,
                'paid': True,
                'context': None,
                'answers': ''
            }
        }

        userdata.update(sample)
    
    wb.close()